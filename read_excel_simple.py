#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
使用 openpyxl 轻量级读取 Excel 文件（避免内存问题）
"""

from openpyxl import load_workbook
import json

def clean_text(text):
    """清理文本"""
    if text is None:
        return ""
    return str(text).strip()

def read_sheet(wb, sheet_name):
    """读取单个工作表"""
    ws = wb[sheet_name]
    data = []
    
    # 读取表头（第一行）
    headers = []
    for cell in ws[1]:
        headers.append(clean_text(cell.value))
    
    print(f"  表头: {headers}")
    
    # 读取数据（从第二行开始）
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = clean_text(value)
        
        # 只保留有数据的行
        if any(row_data.values()):
            data.append(row_data)
    
    return data

def generate_example(item, category):
    """根据分类生成例句"""
    original = item.get('原文表达 (Original)', item.get('原文具体项 (Concrete)', ''))
    target = item.get('题目表达 (Target/Summary)', item.get('题目概括项 (Target/Word)', ''))
    
    # 预定义的一些例句
    examples = {
        "decide to build": {
            "original_sentence": "I decide to build a house.",
            "target_sentence": "I make a ______ to build a house.",
            "answer": "decision"
        },
        "choose": {
            "original_sentence": "I choose the red one.",
            "target_sentence": "I make a ______.",
            "answer": "choice"
        },
        "visit a center": {
            "original_sentence": "I visit a center.",
            "target_sentence": "I pay a ______ to a center.",
            "answer": "visit"
        }
    }
    
    if original in examples:
        return examples[original]
    
    # 默认返回
    return {
        "original_sentence": f"{original}",
        "target_sentence": f"{target}",
        "answer": "[待完善]"
    }

def main():
    excel_path = "E:/credble/读写题/读写题资料.xlsx"
    
    print("开始读取 Excel 文件...")
    wb = load_workbook(excel_path, read_only=True)
    
    all_data = []
    current_id = 1
    
    for sheet_name in wb.sheetnames:
        print(f"\n读取工作表: {sheet_name}")
        rows = read_sheet(wb, sheet_name)
        print(f"  读取到 {len(rows)} 条数据")
        
        for row in rows:
            original = row.get('原文表达 (Original)', row.get('原文具体项 (Concrete)', ''))
            target = row.get('题目表达 (Target/Summary)', row.get('题目概括项 (Target/Word)', ''))
            meaning = row.get('中文释义', row.get('概括范畴', row.get('逻辑转换', '')))
            
            if not original or not target:
                continue
            
            example = generate_example(row, sheet_name)
            
            item = {
                "id": current_id,
                "category": sheet_name,
                "original": original,
                "target": target,
                "meaning": meaning,
                "mastered": False,
                "example": example
            }
            
            all_data.append(item)
            current_id += 1
    
    wb.close()
    
    print(f"\n总共读取 {len(all_data)} 条数据")
    
    # 保存为 JSON
    with open("vocabulary_data.json", 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    
    print("数据已保存到 vocabulary_data.json")
    
    # 统计
    categories = {}
    for item in all_data:
        cat = item['category']
        categories[cat] = categories.get(cat, 0) + 1
    
    print("\n各分类数量:")
    for cat, count in categories.items():
        print(f"  {cat}: {count} 条")

if __name__ == "__main__":
    main()
